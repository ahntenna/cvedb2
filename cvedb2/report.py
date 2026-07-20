import csv
import sys
from datetime import datetime
from pathlib import Path
from sqlite3 import connect, OperationalError
from typing import Dict, Iterable, Iterator, List, Optional, Sequence, Tuple

from .cve import Severity
from .db import CVEdb, DEFAULT_DB_PATH

# (cve_id, feed_rowid) -> last_modified(epoch)
SnapshotKey = Tuple[str, int]
Snapshot = Dict[SnapshotKey, float]

REPORT_COLUMNS = (
    "CVE ID", "Feed", "Published", "Last Modified", "Base Score", "Severity", "Impact Vector", "Description",
)

SQLITE_IN_CHUNK = 500


def take_snapshot(db_path: Path) -> Snapshot:
    if not db_path.exists():
        return {}
    connection = None
    cursor = None
    try:
        connection = connect(str(db_path))
        cursor = connection.cursor()
        try:
            cursor.execute("SELECT id, feed, last_modified FROM cves")
        except OperationalError:
            return {}
        return {(row[0], row[1]): row[2] for row in cursor}
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()


def chunked(items: Sequence, size: int = SQLITE_IN_CHUNK) -> Iterator[Sequence]:
    for i in range(0, len(items), size):
        yield items[i:i + size]


def format_epoch(value) -> str:
    if value is None:
        return ""
    try:
        return datetime.fromtimestamp(float(value)).strftime("%Y-%m-%d %H:%M")
    except (OverflowError, OSError, ValueError):
        return str(value)


def severity_name(value) -> str:
    try:
        return Severity(int(value)).name
    except (KeyError, TypeError, ValueError):
        return ""


def collect_details(db_path: Path, keys: Iterable[SnapshotKey]) -> List[dict]:
    key_set = set(keys)
    if not key_set:
        return []
    ids = sorted({cve_id for cve_id, _ in key_set})
    rows: List[dict] = []
    descriptions: Dict[str, str] = {}
    connection = connect(str(db_path))
    try:
        cursor = connection.cursor()
        for chunk in chunked(ids):
            placeholders = ", ".join(["?"] * len(chunk))
            cursor.execute(
                "SELECT c.id, c.feed, f.name, c.published, c.last_modified, "
                "c.base_score, c.severity, c.impact_vector "
                f"FROM cves c JOIN feeds f ON f.rowid = c.feed WHERE c.id IN ({placeholders})",
                tuple(chunk)
            )
            for cve_id, feed_id, feed_name, published, last_modified, base_score, severity, vector in cursor:
                if (cve_id, feed_id) not in key_set:
                    continue
                rows.append({
                    "cve_id": cve_id,
                    "feed": feed_name,
                    "published": format_epoch(published),
                    "last_modified": format_epoch(last_modified),
                    "base_score": base_score if base_score is not None else "",
                    "severity": severity_name(severity),
                    "impact_vector": vector if vector is not None else "",
                })
            cursor.execute(
                f"SELECT cve, description FROM descriptions WHERE lang = 'en' AND cve IN ({placeholders})",
                tuple(chunk)
            )
            for cve_id, description in cursor:
                descriptions.setdefault(cve_id, description)
    finally:
        connection.close()

    for row in rows:
        row["description"] = descriptions.get(row["cve_id"], "")
    rows.sort(key=lambda r: r["cve_id"], reverse=True)

    return rows


def _row_values(row: dict) -> List:
    return [row["cve_id"], row["feed"], row["published"], row["last_modified"],
            row["base_score"], row["severity"], row["impact_vector"], row["description"]]


def write_csv(path: Path, new_rows: List[dict], modified_rows: List[dict]):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(("Status",) + REPORT_COLUMNS)
        for status, rows in (("NEW", new_rows), ("MODIFIED", modified_rows)):
            for row in rows:
                writer.writerow([status] + _row_values(row))


def write_excel(path: Path, new_rows: List[dict], modified_rows: List[dict], db_path: Path, before_count: int, after_count: int):
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    bold = Font(bold=True)

    summary = wb.active
    summary.title = "Summary"
    summary_rows = (
        ("항목", "값"),
        ("데이터베이스", str(db_path)),
        ("실행 시각", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("업데이트 전 CVE 수", before_count),
        ("업데이트 후 CVE 수", after_count),
        ("신규 추가", len(new_rows)),
        ("내용 변경", len(modified_rows)),
    )
    for row in summary_rows:
        summary.append(row)
    for cell in summary[1]:
        cell.font = bold
    summary.column_dimensions["A"].width = 22
    summary.column_dimensions["B"].width = 60

    column_widths = (18, 8, 17, 17, 11, 10, 45, 100)
    for title, rows in (("New", new_rows), ("Modified", modified_rows)):
        ws = wb.create_sheet(title)
        ws.append(REPORT_COLUMNS)
        for cell in ws[1]:
            cell.font = bold
        for row in rows:
            ws.append(_row_values(row))
        for i, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width
        ws.freeze_panes = "A2"

    wb.save(path)
